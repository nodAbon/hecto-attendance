import importlib.util
import io
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

from openpyxl import load_workbook


TARGET_HEADERS = {
    "ticketNo": ["이용번호"],
    "rideTime": ["탑승일시"],
    "employeeName": ["직원명"],
    "dept": ["부서명"],
    "reason": ["이용사유"],
    "amount": ["결제금액"],
    "settleAmount": ["정산대상금액"],
    "status": ["결제상태"],
    "pickup": ["탑승위치"],
    "dropoff": ["도착위치"],
}


def normalize_value(value):
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if value is None:
        return ""
    return value


def normalize_header(value):
    if value is None:
        return ""
    return str(value).strip()


def find_column_map(headers):
    normalized = [normalize_header(header) for header in headers]
    column_map = {}
    for key, aliases in TARGET_HEADERS.items():
        column_map[key] = None
        for alias in aliases:
            try:
                index = normalized.index(alias)
            except ValueError:
                continue
            column_map[key] = index + 1
            break
    return column_map


def extract_text(ws, row_index, column_index):
    if not column_index:
        return ""
    return normalize_value(ws.cell(row_index, column_index).value)


def main():
    if len(sys.argv) < 3:
        print(json.dumps({"error": "usage: parse_taxi_workbook.py <file> <password>"}))
        sys.exit(1)

    file_path = sys.argv[1]
    password = sys.argv[2]

    pydeps_dir = Path(r"C:\tmp\pydeps")
    if str(pydeps_dir) not in sys.path:
        sys.path.insert(0, str(pydeps_dir))

    msoffcrypto_init = pydeps_dir / "msoffcrypto" / "__init__.py"
    if not msoffcrypto_init.exists():
        print(json.dumps({"error": "msoffcrypto package not found"}))
        sys.exit(2)

    spec = importlib.util.spec_from_file_location(
        "msoffcrypto",
        msoffcrypto_init,
        submodule_search_locations=[str(msoffcrypto_init.parent)],
    )
    if spec is None or spec.loader is None:
        print(json.dumps({"error": "failed to load msoffcrypto package"}))
        sys.exit(2)

    msoffcrypto = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(msoffcrypto)

    with open(file_path, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        decrypted = io.BytesIO()
        office.load_key(password=password)
        office.decrypt(decrypted)

    decrypted.seek(0)
    wb = load_workbook(decrypted, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    column_map = find_column_map(headers)

    rows = []
    for r in range(2, ws.max_row + 1):
      ticket_no = extract_text(ws, r, column_map["ticketNo"])
      ride_time = extract_text(ws, r, column_map["rideTime"])
      employee_name = extract_text(ws, r, column_map["employeeName"])
      dept = extract_text(ws, r, column_map["dept"])
      reason = extract_text(ws, r, column_map["reason"])
      amount = extract_text(ws, r, column_map["amount"])
      settle_amount = extract_text(ws, r, column_map["settleAmount"])
      status = extract_text(ws, r, column_map["status"])
      pickup = extract_text(ws, r, column_map["pickup"])
      dropoff = extract_text(ws, r, column_map["dropoff"])

      has_value = any([
          ticket_no,
          ride_time,
          employee_name,
          dept,
          reason,
          amount,
          settle_amount,
          status,
          pickup,
          dropoff,
      ])
      if not has_value:
          continue

      rows.append({
          "rowIndex": r,
          "ticketNo": str(ticket_no).strip(),
          "rideTime": str(ride_time).strip(),
          "employeeName": str(employee_name).strip(),
          "dept": str(dept).strip(),
          "reason": str(reason).strip(),
          "amount": str(amount).strip(),
          "settleAmount": str(settle_amount).strip(),
          "status": str(status).strip(),
          "pickup": str(pickup).strip(),
          "dropoff": str(dropoff).strip(),
      })

    print(json.dumps({
        "sheetName": ws.title,
        "headers": headers,
        "columnMap": column_map,
        "rows": rows,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
